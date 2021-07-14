import os
import sys
import openpyxl
from openpyxl import load_workbook
import time, calendar,locale
from time import strftime
from datetime import datetime
from datetime import date

#######################################################
#  variables temporelles
locale.setlocale(locale.LC_ALL, '')
a =datetime.now()
y = a.year
m= a.month
d= a.day
datefile = a.strftime('%Y%m%d')
datefile1 = a.strftime('%Y_%m_%d')
moisNm1 = calendar.month_name[m-1]

today = date.today()
format = "%d/%m/%Y"
today = date.today()
date_str = today.strftime(format)

#######################################################
#  fonction de maj cellule EXCEL
def maj_excel_cell(feuille,posrow,poscol,valeur):
    c =feuille.cell(row=posrow, column=poscol)
    c.value = valeur
#######################################################

dir = 'd:/dl/'

wb = load_workbook(filename="modele.xlsx")
ws = wb["DATA"]

list_folder =[]

def find_folders():
    for item in os.listdir():
        if os.path.isfile(item):
            continue
        else:
            yield item

os.chdir(dir)
file_count = 0
ligne = 2
for folder in find_folders():
        print (folder)
               
        #on ecrit le nom du dossier
        maj_excel_cell(ws,ligne,1,str(folder))
       
        # fichiers dans tous les sous repertoires
        file_count = sum(len(files) for _, _, files in os.walk(folder))
        maj_excel_cell(ws,ligne,2,file_count)
        file_count = 0
                
        ligne +=1

wb.save("d:/temp2/Dossiers_utilisateur_" +str(datefile1) + ".xlsx")

file_count = sum(len(files) for _, _, files in os.walk(dir))
print('\nnombre total de fichiers : ',file_count)