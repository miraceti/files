import os
import sys
import openpyxl
from openpyxl import load_workbook
import time, calendar,locale
from time import strftime
from datetime import datetime
from datetime import date

#######################################################
#  variables temporelles
locale.setlocale(locale.LC_ALL, '')
a =datetime.now()
y = a.year
m= a.month
d= a.day
datefile = a.strftime('%Y%m%d')
datefile1 = a.strftime('%Y_%m_%d')
moisNm1 = calendar.month_name[m-1]

today = date.today()
format = "%d/%m/%Y"
today = date.today()
date_str = today.strftime(format)

#######################################################
#  fonction de maj cellule EXCEL
def maj_excel_cell(feuille,posrow,poscol,valeur):
    c =feuille.cell(row=posrow, column=poscol)
    c.value = valeur
#######################################################

dir = 'd:/dl/'

wb = load_workbook(filename="modele.xlsx")

#activation de la feuille
ws = wb["DATA"]

list_folder =[]

def find_folders():
    for item in os.listdir():
        if os.path.isfile(item):
            continue
        else:
            yield item

os.chdir(dir)
ligne = 2
for folder in find_folders():
        print (folder)
        
        list = os.listdir(folder)
        number_files = len(list)
        print("fichiers & dossiers: " + str(number_files))
        
        onlyfiles = next(os.walk(folder))[2]
        print("fichiers uniquement : " + str(len(onlyfiles)))
        
        print('dossier uniquement : ' + str (number_files - len(onlyfiles) ))
        
        print('\n')
        
        maj_excel_cell(ws,ligne,1,str(folder))
        maj_excel_cell(ws,ligne,2,str(number_files))
        maj_excel_cell(ws,ligne,3,str(number_files - len(onlyfiles) ))
        maj_excel_cell(ws,ligne,4,str(len(onlyfiles)))
        
        
        list_folder.append(folder)
        ligne +=1

wb.save("d:/Dossiers_utilisateur_" +str(datefile1) + ".xlsx")

print(list_folder)
print("TERMINE")