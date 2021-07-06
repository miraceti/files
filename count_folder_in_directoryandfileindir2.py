import os
from os import listdir
from os.path import isfile, join
import fnmatch
from os import walk


monrep = "d:\dl"

fichiers = [f for f in listdir(monrep) if isfile(join(monrep, f))]
print(fichiers)
print(len(fnmatch.filter(os.listdir(monrep), "*")))
print(50*'*')

fichiers = [f for f in listdir(monrep) if isfile(join(monrep,f))]
print("p2",fichiers)
print(50*'*')


import os

APP_FOLDER = 'd:/dl/'

totalFiles = 0
totalDir = 0

for base, dirs, files in os.walk(APP_FOLDER):
    print('Searching in : ',base)
    for directories in dirs:
        totalDir += 1
    for Files in files:
        totalFiles += 1


print('Total number of files',totalFiles)
print('Total Number of directories',totalDir)
print('Total:',(totalDir + totalFiles))

print(100*'*')
##################################

dir ='d:/dl/'
list = os.listdir(dir) # dir is your directory path
number_files = len(list)
print (number_files)

# Only files (avoiding subdirectories):



onlyfiles = next(os.walk(dir))[2] #dir is your directory path as string
print (len(onlyfiles))

print('fichier + dossiers : ' + str (number_files))
print('fichiers uniquement : ' + str (len(onlyfiles) ))
print('dossier uniquement : ' + str (number_files - len(onlyfiles) ))

import os 
my_list = os.listdir(dir)
print(my_list)

os.walk(dir)
a = [x[0] for x in os.walk(dir)]
print(a)



#--------*---------*---------*---------*---------*---------*---------*---------*
# Desc: Using 'yield' to get folder names in a directory
#--------*---------*---------*---------*---------*---------*---------*---------*

import os
import sys
import openpyxl
from openpyxl import load_workbook
import time, calendar,locale
from time import strftime
from datetime import datetime
from datetime import date
import datetime as dt
#######################################################
#  variables temporelles
locale.setlocale(locale.LC_ALL, '')
a =datetime.now()
y = a.year
m= a.month
d= a.day
datefile = a.strftime('%Y%m%d')
datefile1 = a.strftime('%Y_%m_%d')
moisNm1 = calendar.month_name[m-1]

today = date.today()
format = "%d/%m/%Y"
today = date.today()
date_str = today.strftime(format)

#######################################################
#  fonction de maj cellule EXCEL
def maj_excel_cell(feuille,posrow,poscol,valeur):
    c =feuille.cell(row=posrow, column=poscol)
    c.value = valeur
#######################################################

wb = load_workbook(filename="modele.xlsx")

#activation de la feuille
ws = wb["DATA"]

list_folder =[]

def find_folders():
    for item in os.listdir():
        if os.path.isfile(item):
            continue
        else:
#                                  # yield folder name     
            yield item


#--------*---------*---------*---------*---------*---------*---------*---------#
# while 1:#                          M A I N L I N E                             #
#--------*---------*---------*---------*---------*---------*---------*---------#
#                                  # set directory
os.chdir(dir)
ligne = 2
for folder in find_folders():
        print (folder)
        
        list = os.listdir(folder) # dir is your directory path
        number_files = len(list)
        print("fichiers & dossiers: " + str(number_files))
        
        onlyfiles = next(os.walk(folder))[2]
        print("fichiers uniquement : " + str(len(onlyfiles)))
        
        print('dossier uniquement : ' + str (number_files - len(onlyfiles) ))
        
        print('\n')
        
        maj_excel_cell(ws,ligne,1,str(folder))
        
        
        list_folder.append(folder)
        ligne +=1
# sys.exit()  

wb.save("Dossiers_utilisateur_" +str(datefile1) + ".xlsx")

print(list_folder)
print("TERMINE")